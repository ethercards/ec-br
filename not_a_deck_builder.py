from openpyxl import load_workbook
from itertools import islice
from collections import OrderedDict
import json



rules_filename = "deckfight2.xlsx"


def count_rows(worksheet):
    row_count = 0
    for row_cells in worksheet.iter_rows():
        if row_cells[0].value is None:
            break
        row_count += 1
    return row_count


def fetch_values(sheet, columns):
    # List to hold dictionaries
    dict_list = []
    # Iterate through non empty rows in worksheet and fetch values into dict
    row_count = count_rows(sheet)
    for row in islice(sheet.values, 1, row_count):
        dict = OrderedDict()
        for value, column in zip(row, columns.keys()):
            # all number is float by default
            if columns[column] == "int":
                if value is not None and not isinstance(value, int):
                    print(
                        f"Invalid int value: ({value}), type: ({type(value)}), column: ({column}) was generated as 0..."
                    )
                    value = None
                dict[column] = int(value) if value is not None else value
            else:
                dict[column] = value
        dict_list.append(dict)
    return dict_list


def load_cards(sheet):
    columns = {
        "id": "str",
        "name": "str",
        "character_type": "str",
        "card_type": "str",
        "combo_sign": "str",
        "attack_action": "str",
        "attack_amount": "str",
        "attack_extra": "str",
        "shield_action": "str",
        "shield_amount": "str",
        "shield_extra": "str",
        "life_action": "str",
        "life_amount": "str",
        "life_extra": "str",
        "crit_action": "str",
        "crit_amount": "str",
        "crit_extra": "str",
        "special": "str",
        "target_opp": "str",
        "target_type": "str",
        "target_subtype": "str",
        "card_timing": "str",
        "card_count": "str",
        "cost": "str"
    }
    cards_data = fetch_values(sheet, columns)
    cards = []
    for cd in cards_data:
        card = {}
        if cd["attack_action"] is not None:
            card["attack"] = {
                "action": cd["attack_action"],
                "amount": int(cd["attack_amount"]),
            }
            if cd["attack_extra"] is not None:
                card["attack"]["extra"] = int(cd["attack_extra"])
        if cd["shield_action"] is not None:
            card["shield"] = {
                "action": cd["shield_action"],
                "amount": int(cd["shield_amount"]),
            }
            if cd["shield_extra"] is not None:
                card["shield"]["extra"] = int(cd["shield_extra"])
        if cd["life_action"] is not None:
            card["life"] = {
                "action": cd["life_action"],
                "amount": int(cd["life_amount"]),
            }
            if cd["life_extra"] is not None:
                card["life"]["extra"] = int(cd["life_extra"])
        if cd["crit_action"] is not None:
            card["crit"] = {
                "action": cd["crit_action"],
                "amount": int(cd["crit_amount"]),
            }
            if cd["crit_extra"] is not None:
                card["crit"]["extra"] = int(cd["crit_extra"])
        if cd["special"] is not None:
            card["special"] = cd["special"]
        if cd["target_opp"] is not None:
            card["target_opp"] = cd["target_opp"]
        if cd["target_type"] is not None:
            card["target_type"] = cd["target_type"]
            if (cd["target_type"] == "card") or (cd["target_type"] == "combo" or (cd["target_type"] == "crit")):
                card["target_subtype"] = cd["target_subtype"]
                card["card_timing"] = int(cd["card_timing"])
                card["card_count"] = int(cd["card_count"])
        card["cost"] = int(cd["cost"])
        card["character_type"] = int(cd["character_type"])
        card["name"] = cd["name"]
        card["card_type"] = cd["card_type"]
        card["combo_sign"] = cd["combo_sign"]
        cards.append(card)

    return cards


workbook = load_workbook(filename=rules_filename)

playing_cards = load_cards(workbook["cards"])

character_type = 3

deck_limit = 20



def create_deck():
    print("You can start building your deck!" )
    print("Please give a valid input of an integer ranging from 0-99 or an x if you finished building your deck")
    deck = []
    deck_cost = 0
    combos=""
    while 1 :
        id = input("Enter a new card id:")
        if len(id)>0:
            if id== "x":
                print("your deck:")
                print(json.dumps(deck))
                return 0
            else:
                id=int(id)
                deck,deck_cost,combos = try_to_add_card_to_deck(id,deck,deck_cost,combos)

def try_to_add_card_to_deck(id,deck, deck_cost,combos):
    card=playing_cards[id]
    if card["character_type"] != character_type:
        print("Invalid card in deck. The card character_type of:", card["character_type"],
              "does not match player character_type of", character_type)
        return deck,deck_cost,combos
    else:
        deck_cost += card["cost"]
    if deck_cost > 200:
        print("invalid deck, deck total cost to high:", deck_cost, "... the maximum allowed is 200")
        deck_cost -= card["cost"]
        return deck,deck_cost,combos
    if len(deck) > deck_limit:
        print("invalid deck, deck has more cards than allowed:", len(deck), "... the maximum allowed is",
            deck_limit)
        deck_cost-=card["cost"]
        return deck,deck_cost,combos

    deck.append({"id":id})
    combos+= card["combo_sign"]
    print("Card added successfully, your total deck cost is:", deck_cost, "cards used so far:", len(deck), "current combo sequence:", combos)
    print("You can add a maximum of",deck_limit-len(deck), "more cards")
    print(json.dumps(deck))
    return deck,deck_cost,combos

create_deck()