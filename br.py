from openpyxl import load_workbook
from itertools import islice
from collections import OrderedDict
import json
import uuid
import random
import jsonpickle
import copy
import os
import sys


rules_filename = "deckfight2.xlsx"
cards_filename = "cards.json"
report_folder_path="reports"
log_reports_folder_path = "log_reports"
players_folder_path='players_info'


# keywords for crating the report
# region
keyword="keyword"
beginning_of_round_report_keyword="beginning_of_round_report"
card_played_keyword="card_played"
phase_started_keyword="phase_started"
phase_ended_keyword="phase_ended"
debuff_applied_keyword="debuff_applied"
card_reducer_debuff_effect_activated_keyword="card_reducer_debuff_effect_activated"
card_reducer_debuff_effectiveness_keyword="card_reducer_debuff_effectiveness"
special_debuff_effect_activated_keyword = "special_debuff_effect_activated"
debuff_card_count_reduced_keyword="debuff_card_count_reduced"
boost_applied_keyword="boost_applied"
combo_found_keyword="combo_found"
combo_boosted_keyword="combo_boosted"
boost_card_count_reduced_keyword="boost_card_count_reduced"
effect_applied_keyword="effect_applied"
defense_effect_activated_keyword="defense_effect_activated"
defense_played_keyword="defense_played"
defense_boosted_keyword="defense_boosted"
defense_applied_keyword="defense_applied"
attack_effect_activated_keyword="attack_effect_activated"
damage_dealt_keyword="damage_deal"
pierce_applied_keyword="pierce_applied"
crit_boosted_keyword="crit_boosted"
attack_played_keyword="attack_played"
attack_boosted_keyword="attack_boosted"
attack_applied_keyword="attack_applied"
end_of_game_report_keyword="end_of_game_report"
end_of_series_report_keyword="end_of_series_report"

# endregion

series_report = []
#TODO keep track of damages dealt
damages_dealt={}
total_healths={}

class Player:

    def __init__(self, card_data, layers, all_combos, all_playing_cards):
        self.params = {
            "combo_group": 0,
            "character_type": 0,
            "crit": 0
        }

        self.data = card_data
        dna = (card_data["layer_image"].split("/")[-1]).split(".")[0]
        self.params["dna"] = dna
        self.params["card_id"] = card_data["id"]
        if dna[0:2] == '04' or dna[0:2] == '004':
            print("goccha")
        self.params["health"] = int(layers[("0" + dna[0:2])]["value"])

        self.params["deck_limit"] = int(layers[("1" + dna[2:4])]["value"])

        self.params["combo_group"] = int(layers[("2" + dna[4:6])]["value"])

        self.params["character_type"] = int(layers[("3" + dna[6:8])]["value"])

        self.params["crit"] = int(layers[("4" + dna[8:10])]["value"])

        self.params["combos"] = []

        self.params["playing_cards"] = []

        for combo in all_combos:
            if combo["type"] == self.params["combo_group"]:
                self.params["combos"].append(combo)

        for playing_card in all_playing_cards:
            if playing_card["character_type"] == self.params["character_type"]:
                self.params["playing_cards"].append(playing_card)

        self.params["deck"] = []
        self.params["deck_cost"] = 0

    def generate_random_deck(self):
        deck = []
        deck_cost = 0
        for card_counter in range(0,self.params["deck_limit"]):
            random_card = random.choice(self.params["playing_cards"])
            # this part is super important (details in documentation)
            card=copy.deepcopy(random_card)
            card["unique_test"]=card_counter
            deck_cost += random_card["cost"]
            deck.append(card)
        self.params["deck"] = deck
        self.params["deck_cost"] = deck_cost
        print("Generated random deck for player:",self.params["card_id"],"deck size:",len(deck),"deck cost:",deck_cost)
        return deck

    def validate_and_assign_deck(self,deck_submitted):
        deck_cost=0
        deck=[]
        for card in deck_submitted:
            if card["character_type"] != self.params["character_type"]:
                self.params["deck"]=[]
                print(self.params["card_id"])
                print("Invalid card in deck. The card character_type of:",card["character_type"],
                      "does not match player character_type of", self.params["character_type"])
                return []
            else:
                deck.append(copy.deepcopy(card))
                deck_cost+=card["cost"]

        if deck_cost > 300:
            self.params["deck"] = []
            print(self.params["card_id"])
            print("invalid deck, deck total cost to high:", deck_cost, "... the maximum allowed is 200")
            return []
        if len(deck) > self.params["deck_limit"]:
            self.params["deck"] = []
            print(self.params["card_id"])
            print("invalid deck, deck has more cards than allowed:", len(deck), "... the maximum allowed is",
                  self.params["deck_limit"])
            return []
        print(self.params["card_id"])
        print("Deck added to the player,deck cost:",deck_cost,"and deck size:",len(deck))
        self.params["deck_cost"]=deck_cost
        self.params["deck"]=deck

    def character(self):
        return self.params


class BattlingPlayerObject:
    def __init__(self, player):
        # initializing permanent components
        self.player_combos = player.params["combos"]
        self.player_dna = player.params["dna"]
        self.player_max_health = player.params["health"]
        self.id=player.params["card_id"]

        # initializing changing components for player1
        self.player_deck = player.params["deck"]
        self.player_health = player.params["health"]
        self.player_shield = 0
        self.player_combo_string = ""
        self.active_boosts = []
        self.combo_effects = []
        self.debuffs = []
        self.player_crit_chance = player.params["crit"]
        self.player_current_deck_size = len(self.player_deck)


class Debuff:

    def __init__(self, neutralizer_card):
        self.neutralizer_card=neutralizer_card
        self.special_debuff = None
        self.card_value_reducer_debuff=None
        self.uuid=str(uuid.uuid4())
        self.card_count = 1
        self.card_timing = 0
        self.evaluate_neutralizer_card(neutralizer_card)
        if "card_timing" in neutralizer_card:
            self.card_timing = int(neutralizer_card["card_timing"])
        if "card_count" in neutralizer_card:
            self.card_count = int(neutralizer_card["card_count"])


    def reduce_card_count(self):
        self.card_count -= 1
        if self.special_debuff is not None:
            self.special_debuff.reduce_card_count()
        if self.card_value_reducer is not None:
            self.card_value_reducer_debuff.reduce_card_count()

    def reduce_card_timing(self):
        self.card_timing-=1
        if self.special_debuff is not None:
            self.special_debuff.reduce_card_timer()
        if self.card_value_reducer_debuff is not None:
            self.card_value_reducer_debuff.reduce_card_timer()

    def evaluate_neutralizer_card(self,neutralizer_card):
        if "special" in neutralizer_card:
            self.special_debuff = SpecialDebuff(neutralizer_card,self.uuid)
        if ("attack" or "shield" or "life" or "crit") in neutralizer_card:
            self.card_value_reducer_debuff = CardValueReducerDebuff(neutralizer_card,self.uuid)


class SpecialDebuff:

    def __init__(self, neutralizer_card,uuid):
        self.neutralizer_card=neutralizer_card
        self.special_description=neutralizer_card["special"]
        self.card_count = 1
        self.uuid = uuid
        self.card_timing = 0
        if "card_timing" in neutralizer_card:
            self.card_timing= int(neutralizer_card["card_timing"])
        if "card_count" in neutralizer_card:
            self.card_count= int(neutralizer_card["card_count"])


    def reduce_card_timer(self):
        self.card_timing-=1

    def reduce_card_count(self):
        self.card_count-=1


class CardValueReducerDebuff:

    def __init__(self,neutralizer_card,uuid):
        self.neutralizer_card=neutralizer_card
        self.card_count = 1
        self.card_timing = 0
        self.uuid = uuid
        if "card_timing" in neutralizer_card:
            self.card_timing = int(neutralizer_card["card_timing"])
        if "card_count" in neutralizer_card:
            self.card_count = int(neutralizer_card["card_count"])
        if "attack" in neutralizer_card:
            self.action=neutralizer_card["attack"]["action"]
            self.amount=neutralizer_card["attack"]["amount"]
        if "life" in neutralizer_card:
            self.action = neutralizer_card["life"]["action"]
            self.amount = neutralizer_card["life"]["amount"]
        if "shield" in neutralizer_card:
            self.action = neutralizer_card["shield"]["action"]
            self.amount = neutralizer_card["shield"]["amount"]
        if "crit" in neutralizer_card:
            self.action = neutralizer_card["crit"]["action"]
            self.amount = neutralizer_card["crit"]["amount"]


    def reduce_card_timer(self):
        self.card_timing -= 1

    def reduce_card_count(self):
        self.card_count -= 1


class ComboEffect:
    def __init__(self, combo_card):
        self.combo_card = combo_card
        self.attack_combo_effect = None
        self.shield_combo_effect = None
        self.life_combo_effect = None
        self.evaluate_combo_card(self.combo_card)

    def evaluate_combo_card(self, combo_card):
        if "attack" in combo_card:
            self.attack_combo_effect = AttackComboEffect(combo_card)
        if "shield" in combo_card:
            self.shield_combo_effect = ShieldComboEffect(combo_card)
        if "life" in combo_card:
            self.life_combo_effect = LifeComboEffect(combo_card)


class AttackComboEffect:
    def __init__(self, combo_card):
        self.action_type = combo_card["attack"]["action"]
        self.min_amount = None
        self.max_amount = None
        if self.action_type == "+":
            self.min_amount = combo_card["attack"]["amount"]
            self.max_amount = combo_card["attack"]["extra"]


class ShieldComboEffect:
    def __init__(self, combo_card):
        self.action_type = combo_card["shield"]["action"]
        self.min_amount = None
        self.max_amount = None
        if self.action_type == "+":
            self.min_amount = combo_card["shield"]["amount"]
            self.max_amount = combo_card["shield"]["extra"]


class LifeComboEffect:
    def __init__(self, combo_card):
        self.action_type = combo_card["life"]["action"]
        self.min_amount = None
        self.max_amount = None
        if self.action_type == "+":
            self.min_amount = combo_card["life"]["amount"]
            self.max_amount = combo_card["life"]["extra"]


class Boost:
    def __init__(self, boost_card):
        self.boost_card = boost_card
        self.unique_id = uuid.uuid4()
        self.attack_boost = None
        self.shield_boost = None
        self.life_boost = None
        self.crit_boost = None
        self.special_boost = None
        self.combo_boost = None
        self.evaluate_boost_card_information(self.boost_card)

    def evaluate_boost_card_information(self, boost_card):
        if boost_card["target_type"] == "combo":
            self.combo_boost = ComboBoost(boost_card, self.unique_id)
        elif "special" in boost_card:
            self.special_boost = SpecialBoost(boost_card, self.unique_id)
        else:
            if "shield" in boost_card:
                self.shield_boost = ShieldBoost(boost_card, self.unique_id)
            if "life" in boost_card:
                self.life_boost = LifeBoost(boost_card, self.unique_id)
            if "attack" in boost_card:
                self.attack_boost = AttackBoost(boost_card, self.unique_id)
            if "crit" in boost_card:
                self.crit_boost = CritBoost(boost_card, self.unique_id)


class ComboBoost:
    def __init__(self, boost_card, id):
        self.boost_card = boost_card
        self.unique_id = id
        self.target_type = boost_card["target_type"]

        self.card_timing = int(boost_card["card_timing"])
        self.card_count = int(boost_card["card_count"])

        if "attack" in boost_card:
            self.action_type = boost_card["attack"]["action"]
            self.amount = int(boost_card["attack"]["amount"])

        if "crit" in boost_card:
            self.action_type = boost_card["crit"]["action"]
            self.amount = int(boost_card["crit"]["amount"])

        if "shield" in boost_card:
            self.action_type = boost_card["shield"]["action"]
            self.amount = int(boost_card["shield"]["amount"])

        if "life" in boost_card:
            self.action_type = boost_card["life"]["action"]
            self.amount = int(boost_card["life"]["amount"])


class AttackBoost:
    def __init__(self, boost_card, id):
        self.boost_card = boost_card
        self.unique_id = id
        self.target_type = boost_card["target_type"]
        self.target_opp = ["target_opp"]
        self.action_type = boost_card["attack"]["action"]
        self.amount = int(boost_card["attack"]["amount"])
        self.extra=None
        if "extra" in boost_card["attack"]:
            self.extra=int(boost_card["attack"]["extra"])

        self.card_timing = int(boost_card["card_timing"])
        self.card_count = int(boost_card["card_count"])


class LifeBoost:
    def __init__(self, boost_card, id):
        self.boost_card = boost_card
        self.unique_id = id
        self.target_type = boost_card["target_type"]
        self.target_opp = boost_card["target_opp"]
        self.action_type = boost_card["life"]["action"]
        self.amount = int(boost_card["life"]["amount"])
        self.extra=None
        if "extra" in boost_card["life"]:
            self.extra = int(boost_card["life"]["extra"])

        self.card_timing = int(boost_card["card_timing"])
        self.card_count = int(boost_card["card_count"])


class ShieldBoost:
    def __init__(self, boost_card, id):
        self.boost_card = boost_card
        self.unique_id = id
        self.target_type = boost_card["target_type"]
        self.target_opp = boost_card["target_opp"]
        self.action_type = boost_card["shield"]["action"]
        self.amount = int(boost_card["shield"]["amount"])
        self.extra=None

        self.card_timing = int(boost_card["card_timing"])
        self.card_count = int(boost_card["card_count"])

        if "extra" in boost_card["shield"]:
            self.extra = int(boost_card["shield"]["amount"])


class CritBoost:
    def __init__(self, boost_card, id):
        self.boost_card = boost_card
        self.unique_id = id
        self.target_type = boost_card["target_type"]
        self.target_opp = boost_card["target_opp"]
        self.action_type = boost_card["crit"]["action"]
        self.amount = int(boost_card["crit"]["amount"])

        self.card_timing = int(boost_card["card_timing"])
        self.card_count = int(boost_card["card_count"])


class SpecialBoost:
    def __init__(self, boost_card, id):
        self.boost_card = boost_card
        self.unique_id = id
        self.target_type = boost_card["target_type"]
        self.target_opp = boost_card["target_opp"]
        self.special = boost_card["special"]

        self.card_timing = int(boost_card["card_timing"])
        self.card_count = int(boost_card["card_count"])
#############################

def count_rows(worksheet):
    row_count = 0
    for row_cells in worksheet.iter_rows():
        if row_cells[0].value is None:
            break
        row_count += 1
    return row_count


def fetch_values(sheet, columns):
    # List to hold dictionaries
    dict_list = []
    # Iterate through non empty rows in worksheet and fetch values into dict
    row_count = count_rows(sheet)
    for row in islice(sheet.values, 1, row_count):
        dict = OrderedDict()
        for value, column in zip(row, columns.keys()):
            # all number is float by default
            if columns[column] == "int":
                if value is not None and not isinstance(value, int):
                    print(
                        f"Invalid int value: ({value}), type: ({type(value)}), column: ({column}) was generated as 0..."
                    )
                    value = None
                dict[column] = int(value) if value is not None else value
            else:
                dict[column] = value
        dict_list.append(dict)
    return dict_list


def load_layers(sheet):
    columns = {
        "seq": "str",
        "code": "str",
        "Type": "str",
        "value": "str"
    }
    ls = fetch_values(sheet, columns)
    layers = {}
    for layer in ls:
        k = str(int(layer["seq"]))
        c = layer["code"]
        if isinstance(c, int):
            c = int(c)
        c = str(c)
        k = k + c
        layers[k] = layer

    return (layers)


def load_cards(sheet):
    columns = {
        "id": "str",
        "name": "str",
        "character_type": "str",
        "card_type": "str",
        "combo_sign": "str",
        "attack_action": "str",
        "attack_amount": "str",
        "attack_extra": "str",
        "shield_action": "str",
        "shield_amount": "str",
        "shield_extra": "str",
        "life_action": "str",
        "life_amount": "str",
        "life_extra": "str",
        "crit_action": "str",
        "crit_amount": "str",
        "crit_extra": "str",
        "special": "str",
        "target_opp": "str",
        "target_type": "str",
        "target_subtype": "str",
        "card_timing": "str",
        "card_count": "str",
        "cost": "str"
    }
    cards_data = fetch_values(sheet, columns)
    cards = []
    for cd in cards_data:
        card = {}
        if cd["attack_action"] is not None:
            card["attack"] = {
                "action": cd["attack_action"],
                "amount": int(cd["attack_amount"]),
            }
            if cd["attack_extra"] is not None:
                card["attack"]["extra"] = int(cd["attack_extra"])
        if cd["shield_action"] is not None:
            card["shield"] = {
                "action": cd["shield_action"],
                "amount": int(cd["shield_amount"]),
            }
            if cd["shield_extra"] is not None:
                card["shield"]["extra"] = int(cd["shield_extra"])
        if cd["life_action"] is not None:
            card["life"] = {
                "action": cd["life_action"],
                "amount": int(cd["life_amount"]),
            }
            if cd["life_extra"] is not None:
                card["life"]["extra"] = int(cd["life_extra"])
        if cd["crit_action"] is not None:
            card["crit"] = {
                "action": cd["crit_action"],
                "amount": int(cd["crit_amount"]),
            }
            if cd["crit_extra"] is not None:
                card["crit"]["extra"] = int(cd["crit_extra"])
        if cd["special"] is not None:
            card["special"] = cd["special"]
        if cd["target_opp"] is not None:
            card["target_opp"] = cd["target_opp"]
        if cd["target_type"] is not None:
            card["target_type"] = cd["target_type"]
            if (cd["target_type"] == "card") or (cd["target_type"] == "combo" or (cd["target_type"] == "crit")):
                card["target_subtype"] = cd["target_subtype"]
                card["card_timing"] = int(cd["card_timing"])
                card["card_count"] = int(cd["card_count"])
        card["cost"] = int(cd["cost"])
        card["character_type"] = int(cd["character_type"])
        card["name"] = cd["name"]
        card["card_type"] = cd["card_type"]
        card["combo_sign"] = cd["combo_sign"]
        cards.append(card)

    return cards


def load_combos(sheet):
    columns = {
        "type": "str",
        "combo": "str",
        "attack_action": "str",
        "attack_amount": "str",
        "attack_extra": "str",
        "shield_action": "str",
        "shield_amount": "str",
        "shield_extra": "str",
        "life_action": "str",
        "life_amount": "str",
        "life_extra": "str",
        "crit_action": "str",
        "crit_amt": "str",
        "crit_extra": "str",
        "special": "str",
        "target_opp": "str",
        "target_type": "str",
        "card_timing": "str",
        "card_count": "str"

    }
    cs = fetch_values(sheet, columns)
    combos = []
    for c in cs:
        combo = {}
        combo["type"] = int(c["type"])

        # attack
        if c["attack_action"] is not None:
            combo["attack"] = {
                "action": c["attack_action"],
                "amount": int(c["attack_amount"]),
            }
            if c["attack_extra"] is not None:
                combo["attack"]["extra"] = int(c["attack_extra"])

        # shield
        if c["shield_action"] is not None:
            combo["shield"] = {
                "action": c["shield_action"],
                "amount": int(c["shield_amount"]),
            }
            if c["shield_extra"] is not None:
                combo["shield"]["extra"] = int(c["shield_extra"])

        # life
        if c["life_action"] is not None:
            combo["life"] = {
                "action": c["life_action"],
                "amount": int(c["life_amount"]),
            }
            if c["life_extra"] is not None:
                combo["life"]["extra"] = int(c["life_extra"])

        # crit
        if c["crit_action"] is not None:
            combo["crit"] = {
                "action": c["crit_action"],
                "amount": int(c["crit_amt"]),
            }
        combo["combo_code"] = c["combo"]
        combo["target_opp"] = c["target_opp"]
        combo["target_type"] = c["target_type"]
        combo["card_timing"] = c["card_timing"]
        combo["card_count"] = c["card_count"]

        combos.append(combo)
    return combos


def add_to_report(data):
    series_report.append(data)


def create_phase_started_report(phase_name):
    data = {
        keyword:phase_started_keyword,
        "phase_name": phase_name
    }
    return  data


def create_phase_ended_report(phase_name):
    data = {
        keyword:phase_ended_keyword,
        "phase_name": phase_name
    }
    return  data


def create_card_played_report(battling_player_id,card_played):
    data ={
        keyword:card_played_keyword,
        "player_id": battling_player_id,
        card_played_keyword: card_played
    }
    return  data


def create_card_reducer_debuff_effectiveness_report(defending_player_id, effectiveness):
    data={
        keyword:card_reducer_debuff_effectiveness_keyword,
        "defending_player_id":defending_player_id,
        "effectiveness":effectiveness
    }
    return  data

def create_debuff_applied_report(attacking_player_id,defending_player_id, debuff):
    data = {
        keyword : debuff_applied_keyword,
        "defending_player_id": defending_player_id,
        "attacking_player_id":attacking_player_id,
        "debuff" : debuff
    }

    return  data


def create_card_reducer_debuff_effect_activated_report(defending_player_id,parameter, action_type, old_values, new_values, amount):
    data = {
        keyword:card_reducer_debuff_effect_activated_keyword,
        "defending_player_id":defending_player_id,
        "parameter": parameter,
        "action_type":action_type,
        "amount":amount,
        "old_values" : old_values,
        "new_values" : new_values
    }

    return  data


def create_debuff_card_count_reduced_report(message, card_count_left):
    data ={
        keyword : debuff_card_count_reduced_keyword,
        "message": message,
        "card_count_left": card_count_left
    }
    return  data

def create_special_debuff_activated_report(defending_player_id,message):
    data = {
        keyword : special_debuff_effect_activated_keyword,
        "defending_player_id":defending_player_id,
        "message": message
    }
    return  data


def create_boost_applied_report(active_player_id,boost):
    data={
        keyword: boost_applied_keyword,
        "active_player_id": active_player_id,
        "boost": boost
    }

    return  data

def create_combo_found_report(active_player_id, combo_string,combo_level, combo):
    data ={
        keyword:combo_found_keyword,
        "active_player_id":active_player_id,
        "como_string": combo_string,
        "combo_level": combo_level,
        "combo": combo
    }
    return  data

def create_combo_boosted_report(active_player_id,parameter,action_type,old_values,new_values,amount ):
    data = {
        keyword : combo_boosted_keyword,
        "active_player_id":active_player_id,
        "parameter": parameter,
        "action_type":action_type,
        "amount": amount,
        "old_values": old_values,
        "new_values": new_values
    }
    return  data


def create_boost_card_count_reduced_keyword(message,card_count_left):
    data = {
        keyword: boost_card_count_reduced_keyword,
        "message": message,
        "card_count_left": card_count_left
    }
    return data

def create_effect_applied_report(active_player_id, effect):
    data= {
        keyword: effect_applied_keyword,
        "active_player_id" : active_player_id,
        "effect" : effect
    }
    return data

def create_defense_effect_activated_report(defending_player_id,defense_type, gained_amount,new_value):
    data = {
        keyword : defense_effect_activated_keyword,
        "defending_player_id":defending_player_id,
        "defense_type":defense_type,
        "gained_amount":gained_amount,
        "new_value": new_value
    }
    return  data


def create_defense_played_report(defending_player_id, defense_type, min_value, max_value, final_value):
    data = {
        keyword : defense_played_keyword,
        "defending_player_id": defending_player_id,
        "defense_type": defense_type,
        "min_value": min_value,
        "max_value": max_value,
        "final_value": final_value,
    }
    return  data


def create_defense_boosted_report(defending_player_id, defense_type, boost_amount, boost_type, old_value, new_value):
    data = {
        keyword : defense_boosted_keyword,
        "defending_player_id": defending_player_id,
        "defense_type": defense_type,
        "boost_amount": boost_amount,
        "boost_type": boost_type,
        "old_value": old_value,
        "new_value": new_value
    }
    return  data


def create_defense_applied_report(defending_player_id, defense_type, amount, new_health, new_shield):
    data = {
        keyword: defense_applied_keyword,
        "defending_player_id": defending_player_id,
        "defense_type":defense_type,
        "amount": amount,
        "new_health": new_health,
        "new_shield": new_shield
    }

    return data


def create_attack_effect_activated_report(attacking_player_id, min_value, max_value, final_value):
    data ={
        keyword: attack_effect_activated_keyword,
        "attacking_player_id": attacking_player_id,
        "min_value": min_value,
        "max_value": max_value,
        "final_value": final_value
    }
    return data


def create_damage_dealt_report(attacking_player_id,defending_player_id, damage_blocked, damage_dealt,new_health, new_shield):
    data ={
        keyword : damage_dealt_keyword,
        "attacking_player_id":attacking_player_id,
        "defending_player_id":defending_player_id,
        "damage_blocked": damage_blocked,
        "damage_dealt": damage_dealt,
        "new_health":new_health,
        "new_shield": new_shield
    }
    return  data


def create_pierce_applied_report(attacking_player_id, pierce_applied):
    data= {
        keyword:pierce_applied_keyword,
        "attacking_player":attacking_player_id,
        "pierce_applied": pierce_applied
    }
    return data


def create_crit_boosted(attacking_player_id,action_type, amount, old_value, new_value):
    data = {
        keyword:crit_boosted_keyword,
        "attacking_player_id":attacking_player_id,
        "amount": amount,
        "action_type":action_type,
        "old_value": old_value,
        "new_value": new_value
    }
    return data


def create_attack_played_report(attacking_player_id, min_attack, max_attack, crit_chance, was_crit, final_damage):
    data = {
        keyword: attack_played_keyword,
        "attacking_player_id": attacking_player_id,
        "min_attack": min_attack,
        "max_attack": max_attack,
        "crit_chance": crit_chance,
        "was_crit":was_crit,
        "final_damage": final_damage
    }

    return data


def create_attack_boosted_report(attacking_player_id, boost_type, boost_amount, old_value, new_value):
    data = {
        keyword: attack_boosted_keyword,
        "attacking_player_id": attacking_player_id,
        "boost_type":boost_type,
        "boost_amount": boost_amount,
        "old_value": old_value,
        "new_value": new_value
    }
    return data


def create_attack_applied(attacking_player_id, final_damage, is_piercing):
    data = {
        keyword:attack_applied_keyword,
        "attacking_player_id":attacking_player_id,
        "final_damage": final_damage,
        "is_piercing":is_piercing
    }
    return data


def create_beginning_of_round_report(round_counter,battling_player1,battling_player2):
    data ={
        "key_word": beginning_of_round_report_keyword,
        "round":round_counter,
        "players_info":[{
            "player_id":battling_player1.id,
            "shield": battling_player1.player_shield,
            "health": battling_player1.player_health,
            "active_boosts": len(battling_player1.active_boosts),
            "combo_effects": len(battling_player1.combo_effects),
            "debuffs": len(battling_player1.debuffs),
            "crit_chace": battling_player1.player_crit_chance,
            "deck_size": len(battling_player1.player_deck),
            "combo_string" : battling_player1.player_combo_string
        },{
            "player_id": battling_player2.id,
            "shield": battling_player2.player_shield,
            "health": battling_player2.player_health,
            "active_boosts": len(battling_player2.active_boosts),
            "combo_effects": len(battling_player2.combo_effects),
            "debuffs": len(battling_player2.debuffs),
            "crit_chace": battling_player2.player_crit_chance,
            "deck_size": len(battling_player2.player_deck),
            "combo_string": battling_player2.player_combo_string

        }]
    }
    return  data


def create_end_of_game_report(winner_id, loser_id):
    data={
        keyword:end_of_game_report_keyword,
        "winner_id": winner_id,
        "loser_id" : loser_id
    }
    return  data

def create_end_of_series_report(winner_id, loser_id,total_damages_dealt, final_hps,score):
    data={
        keyword:end_of_series_report_keyword,
        "winner_id":winner_id,
        "loser_id": loser_id,
        "total_damages_dealt":total_damages_dealt,
        "final_hps":final_hps,
        "score":score
    }
    return data


def battle(player1, player2):
    player1_copy=copy.deepcopy(player1)
    player2_copy=copy.deepcopy(player2)
    print("****************************************")
    print("NEW BATTLE")
    print("****************************************")
    round_counter = 1
    battling_player1 = BattlingPlayerObject(player1_copy)
    battling_player2 = BattlingPlayerObject(player2_copy)

    print_overall_info(battling_player1)
    print_overall_info(battling_player2)
    data= create_beginning_of_round_report(round_counter,battling_player1,battling_player2)
    add_to_report(data)
    print("+++++++++++++++++++++")
    print("Round ", round_counter, ": ")
    print("+++++++++++++++++++++")
    round_counter += 1
    while evaluate_round(battling_player1, battling_player2):
        data = create_beginning_of_round_report(round_counter, battling_player1, battling_player2)
        add_to_report(data)
        print("+++++++++++++++++++++")
        print("Round ", round_counter, ": ")
        print("+++++++++++++++++++++")
        print("----------------------------")
        print("Beggining of round player state:")
        print("----------------------------")
        print_overall_info(battling_player1)
        print_overall_info(battling_player2)
        print("----------------------------")
        print("Round evaluation begins:")
        print("----------------------------")
        round_counter += 1
    return determine_winner(battling_player1, battling_player2)


def print_overall_info(battling_player):
    print("player dna:", battling_player.player_dna)
    print(battling_player.player_dna, "health:", battling_player.player_health, "shield:",
          battling_player.player_shield, "crit:", battling_player.player_crit_chance)
    print(battling_player.player_dna, "deck size:", len(battling_player.player_deck),"current combo string:",battling_player.player_combo_string, "active boost:",
          len(battling_player.active_boosts), "active debuffs:", len(battling_player.debuffs))


def evaluate_round(battling_player1, battling_player2):

    # game over condition

    battling_player1_card_to_play = try_to_play_card(battling_player1)
    data = create_card_played_report(battling_player1.id,battling_player1_card_to_play)
    add_to_report(data)

    battling_player2_card_to_play = try_to_play_card(battling_player2)
    data = create_card_played_report(battling_player2.id, battling_player2_card_to_play)
    add_to_report(data)
    
    evaluate_cards(battling_player1, battling_player1_card_to_play, battling_player2, battling_player2_card_to_play)

    battling_player1.combo_effects = []
    battling_player2.combo_effects = []
    print("----------------------------")
    print("End of round player state:")
    print("----------------------------")

    print_overall_info(battling_player1)
    print_overall_info(battling_player2)
    print("+++++++++++++++++++++")
    print("Round OVER")
    print("+++++++++++++++++++++")

    if (battling_player1.player_health < 1 or battling_player2.player_health < 1) or \
            (battling_player1.player_current_deck_size == 0 and battling_player2.player_current_deck_size == 0):
        return False

    return True


def try_to_play_card(battling_player):
    if len(battling_player.player_deck) != 0:
        card_to_play = battling_player.player_deck.pop(0)
        battling_player.player_current_deck_size = len(battling_player.player_deck)
        print_card_info(battling_player,card_to_play)
    else:
        card_to_play = None
        print(battling_player.player_dna,"has no more cards")

    return card_to_play


def print_card_info(battling_player,card_to_play):
    print(battling_player.player_dna, " plays ", card_to_play)


def evaluate_cards(battling_player1, battling_player1_card_to_play, battling_player2, battling_player2_card_to_play):
    # technically revealing the two cards and adding their combo signs to the string
    if battling_player1_card_to_play is not None:
        battling_player1.player_combo_string += battling_player1_card_to_play["combo_sign"]
    if battling_player2_card_to_play is not None:
        battling_player2.player_combo_string += battling_player2_card_to_play["combo_sign"]

    # evaluate neutralizer phase
    (battling_player1,battling_player1_card_to_play,battling_player2,battling_player2_card_to_play) \
        = evaluate_neutralizer_phase(battling_player1,battling_player1_card_to_play,
                                                                     battling_player2,battling_player2_card_to_play)

    # evaluate boost phase
    (battling_player1, battling_player2) = evaluate_boost_phase(battling_player1, battling_player1_card_to_play,
                                                                battling_player2, battling_player2_card_to_play)

    # evaluate combo phase
    (battling_player1, battling_player2) = evaluate_combo_phase(battling_player1, battling_player2)

    # evaluate defense phase
    (battling_player1, battling_player2) = evaluate_defense_phase(battling_player1, battling_player1_card_to_play,
                                                                  battling_player2, battling_player2_card_to_play)

    # evaluate attack phase
    (battling_player1, battling_player2) = evaluate_attack_phase(battling_player1, battling_player1_card_to_play,
                                                                 battling_player2, battling_player2_card_to_play)

    return battling_player1, battling_player2


def evaluate_neutralizer_phase(battling_player1, battling_player1_card_to_play,
                         battling_player2, battling_player2_card_to_play):
    print("----------------------------")
    print("Neutralizer phase started")
    print("----------------------------")
    data= create_phase_started_report("neutralizer_phase")
    add_to_report(data)
    battling_player1,battling_player1_card_to_play,battling_player2,battling_player2_card_to_play \
        = evaluate_neutralizer_phase_for_player(
        battling_player1,battling_player1_card_to_play,battling_player2,battling_player2_card_to_play)

    battling_player2,battling_player2_card_to_play,battling_player1,battling_player1_card_to_play \
        = evaluate_neutralizer_phase_for_player(
        battling_player2,battling_player2_card_to_play,battling_player1,battling_player1_card_to_play)

    data = create_phase_ended_report("neutralizer_phase")
    add_to_report(data)
    print("----------------------------")
    print("Neutralizer phase ended")
    print("----------------------------")

    return battling_player1,battling_player1_card_to_play,battling_player2,battling_player2_card_to_play


def evaluate_neutralizer_phase_for_player(attacking_player, attacking_card_to_play, defending_player, defending_card_to_play):

    if attacking_card_to_play is not None:
        if attacking_card_to_play["combo_sign"] == "N":
            attacking_player,defending_player=evaluate_neutralizer_card(attacking_player, attacking_card_to_play,
                                                                        defending_player)

    defending_player,defending_card_to_play = apply_neutralizer_debuffs(defending_player,defending_card_to_play)

    return attacking_player,attacking_card_to_play,defending_player,defending_card_to_play


def evaluate_neutralizer_card(attacking_player,card_to_play, defending_player):

    debuff = Debuff(card_to_play)
    defending_player.debuffs.append(debuff)
    data= create_debuff_applied_report(attacking_player.id,defending_player.id,card_to_play)
    add_to_report(data)
    print (attacking_player.player_dna,debuff.neutralizer_card,"debuff applied on", defending_player.player_dna)
    return attacking_player,defending_player


def apply_neutralizer_debuffs(defending_player,defending_card_to_play):
    # todo nezzuk ezt meg at Szilagyival
    if defending_card_to_play is not None :
        if len(defending_player.debuffs) > 0:
            for debuff in defending_player.debuffs:
                if debuff.card_timing > 0:
                    debuff.reduce_card_timing()
                else:
                    if debuff.special_debuff is not None and debuff.card_count>0:
                        special_description=debuff.special_debuff.special_description
                        if special_description == "swap":
                            apply_deck_manipulator_debuff(defending_player,special_description)
                        if special_description == "to the bottom":
                            apply_deck_manipulator_debuff(defending_player,special_description)
                        if special_description == "combo sign cancel":
                            apply_combo_sign_cancel(defending_player)
                    if debuff.card_value_reducer_debuff is not None and debuff.card_count>0:
                        apply_card_reducer_debuff(defending_player,debuff.card_value_reducer_debuff,defending_card_to_play)
                    remove_debuff_counter(defending_player,debuff.uuid)
    return defending_player,defending_card_to_play


def remove_debuff_counter(defending_player,debuff_uuid):
    debuff_to_remove_index = None
    for index in range(len(defending_player.debuffs)):
        if defending_player.debuffs[index].uuid == debuff_uuid:
            if int(defending_player.debuffs[index].card_count) == 1:
                debuff_to_remove_index = index
                data = create_debuff_card_count_reduced_report("Debuff has no more charges",0)
                add_to_report(data)
                print("debuff with id:", defending_player.debuffs[index].neutralizer_card, "has no more charges, removing it")
            else:
                new_card_count = int(defending_player.debuffs[index].card_count) - 1
                defending_player.debuffs[index].card_count = new_card_count
                data = create_debuff_card_count_reduced_report("Debuff has more than 1 charge, removing 1 from it, new charges left:", new_card_count)
                add_to_report(data)
                print("debuff with id:", defending_player.debuffs[index].neutralizer_card, "has more than 1 charge, removing 1 from it, new charges left:",
                      defending_player.debuffs[index].card_count)

    if debuff_to_remove_index is not None:
        defending_player.debuffs.pop(debuff_to_remove_index)

    return defending_player


def apply_card_reducer_debuff(defending_player, card_value_reducer_debuff, card_played):
    ineffective_debuff = True
    if card_played["combo_sign"] != "N":
        if card_played["combo_sign"] == "B":
            if "attack" in card_played:
               ineffective_debuff=apply_card_reducer_debuff_on_boost_card_with_parameter(defending_player,card_value_reducer_debuff,card_played,"attack")
            if "shield" in card_played:
                ineffective_debuff=apply_card_reducer_debuff_on_boost_card_with_parameter(defending_player,card_value_reducer_debuff,card_played,"shield")
            if "life" in card_played:
                ineffective_debuff=apply_card_reducer_debuff_on_boost_card_with_parameter(defending_player,card_value_reducer_debuff,card_played,"life")
            if "crit" in card_played:
                ineffective_debuff = apply_card_reducer_debuff_on_boost_card_with_parameter(defending_player,
                                                                                            card_value_reducer_debuff,
                                                                                            card_played, "crit")
        else:
            if "attack" in card_played:
                ineffective_debuff=apply_card_reducer_debuff_on_non_boost_card_with_parameter(defending_player,card_value_reducer_debuff,card_played,"attack")
            if "shield" in card_played:
                ineffective_debuff=apply_card_reducer_debuff_on_non_boost_card_with_parameter(defending_player,card_value_reducer_debuff,card_played,"shield")
            if "life" in card_played:
                ineffective_debuff = apply_card_reducer_debuff_on_non_boost_card_with_parameter(defending_player,card_value_reducer_debuff,card_played, "life")
            if "crit" in card_played:
                ineffective_debuff=apply_card_reducer_debuff_on_non_boost_card_with_parameter(defending_player,card_value_reducer_debuff,card_played,"crit")

    if ineffective_debuff:
        data= create_card_reducer_debuff_effectiveness_report(defending_player.id,"useless")
        add_to_report(data)
        print("Useless card reducer debuff applied on", defending_player.player_dna)
    else:
        data = create_card_reducer_debuff_effectiveness_report(defending_player.id, "effective")
        add_to_report(data)
        print("Effective card reducer debuff applied on on", defending_player.player_dna)

def apply_card_reducer_debuff_on_boost_card_with_parameter(defending_player,card_value_reducer_debuff,card_played,parameter):
    min_value = card_played[parameter]["amount"]
    action_type = card_played[parameter]["action"]
    if card_value_reducer_debuff.action == "=":
        card_played[parameter]["amount"] = card_value_reducer_debuff.amount
    if card_value_reducer_debuff.action == "-":
        card_played[parameter]["amount"] -= card_value_reducer_debuff.amount
    if card_played[parameter]["action"] == "x":
        if card_played[parameter]["amount"] < 1:
            card_played[parameter]["amount"] = 1
    elif card_played[parameter]["action"] == "+":
        if card_played[parameter]["amount"] < 0:
            card_played[parameter]["amount"] = 0
    new_value = card_played[parameter]["amount"]
    
    data = create_card_reducer_debuff_effect_activated_report(defending_player.id,parameter,card_value_reducer_debuff.action,min_value,new_value,card_value_reducer_debuff.amount)
    add_to_report(data)
    print(defending_player.player_dna, "card reducer debuff applied on",parameter," boost card, from", action_type, min_value,
          "to", new_value)
    return False


def apply_card_reducer_debuff_on_non_boost_card_with_parameter(defending_player,card_value_reducer_debuff,card_played,parameter):
    min_value = card_played[parameter]["amount"]
    if "extra" in card_played[parameter]:
        max_value = card_played[parameter]["extra"]
    action_type = card_played[parameter]["action"]
    if card_value_reducer_debuff.action == "=":
        card_played[parameter]["amount"] = card_value_reducer_debuff.amount
        if "extra" in card_played[parameter]:
            card_played[parameter]["extra"] = card_value_reducer_debuff.amount
    if card_value_reducer_debuff.action == "-":
        card_played[parameter]["amount"] -= card_value_reducer_debuff.amount
        if "extra" in card_played[parameter]:
            card_played[parameter]["extra"] -= card_value_reducer_debuff.amount
        if card_played[parameter]["amount"] < 0:
            card_played[parameter]["amount"] = 0
        if card_played[parameter]["extra"] < 0:
            if "extra" in card_played[parameter]:
                card_played[parameter]["extra"] = 0
    if "extra" in card_played[parameter]:
        old_values=min_value, "-",max_value
        new_values=card_played[parameter]["amount"], "-", card_played[parameter]["extra"]
        data = create_card_reducer_debuff_effect_activated_report(defending_player.id, parameter, card_value_reducer_debuff.action,
                                                                  old_values, new_values,
                                                                  card_value_reducer_debuff.amount)
        add_to_report(data)
        print(defending_player.player_dna, "card reducer debuff applied on",parameter,"card, from", card_value_reducer_debuff.action, min_value, "-",
              max_value,
              "to", card_played[parameter]["amount"], "-", card_played[parameter]["extra"])
    else:
        data = create_card_reducer_debuff_effect_activated_report(defending_player.id, parameter, action_type,
                                                                  min_value, new_value,
                                                                  card_value_reducer_debuff.amount)
        add_to_report(data)
        print(defending_player.player_dna, "card reducer debuff applied on", parameter, "card, from", card_value_reducer_debuff.action,
              min_value,"to", card_played[parameter]["amount"])

    return False


def apply_combo_sign_cancel(defending_player):
    previous_combo_string=defending_player.player_combo_string
    #defending_player.player_combo_string= defending_player.player_combo_string[:-1]
    defending_player.player_combo_string=""
    data = create_special_debuff_activated_report(defending_player.id,"Combo string reset")
    add_to_report(data)
    print(defending_player.player_dna,"Combo string changed from:",previous_combo_string,"to",defending_player.player_combo_string)



def apply_deck_manipulator_debuff(defending_player,debuff_description):
    if len(defending_player.player_deck) > 1:
        if debuff_description == "swap":
            first_card =  defending_player.player_deck[0]
            second_card =defending_player.player_deck[1]
            defending_player.player_deck[0]=second_card
            defending_player.player_deck[1]=first_card
            data= create_special_debuff_activated_report(defending_player.id,"First two cards swapped")
            add_to_report(data)
            print(defending_player.player_dna, "first two cards swapped")
        if debuff_description == "to the bottom":
            first_card=defending_player.player_deck[0]
            defending_player.player_deck.pop(0)
            defending_player.player_deck.append(first_card)
            data = create_special_debuff_activated_report(defending_player.id, "First card sent to botoom")
            add_to_report(data)
            print (defending_player.player_dna,"first card sent to bottom")



def evaluate_boost_phase(battling_player1, battling_player1_card_to_play,
                         battling_player2, battling_player2_card_to_play):
    print("----------------------------")
    print("Boost phase started")
    print("----------------------------")
    data = create_phase_started_report("boost_phase")
    add_to_report(data)
    evaluate_boost_phase_for_player(battling_player1,battling_player1_card_to_play)

    evaluate_boost_phase_for_player(battling_player2,battling_player2_card_to_play)
    data = create_phase_ended_report("boost_phase")
    add_to_report(data)
    print("----------------------------")
    print("Boost phase ended")
    print("----------------------------")

    return battling_player1, battling_player2


def evaluate_boost_phase_for_player(battling_player,battling_player_card_to_play):
    if battling_player_card_to_play is not None:
        if battling_player_card_to_play["combo_sign"] == "B":
            boost = Boost(battling_player_card_to_play)
            battling_player.active_boosts.append(boost)
            data = create_boost_applied_report(battling_player.id,battling_player_card_to_play)
            add_to_report(data)
            print(battling_player.player_dna,"added the following boost",boost.boost_card)


def remove_boost_counter(active_player, boost_id):
    boost_to_remove_index = None
    for index in range(len(active_player.active_boosts)):
        if active_player.active_boosts[index].unique_id == boost_id:
            if int(active_player.active_boosts[index].boost_card["card_count"]) == 1:
                boost_to_remove_index = index
                data= create_boost_card_count_reduced_keyword("Boost has no more charges, removing it",0)
                add_to_report(data)
                print("boost card:", active_player.active_boosts[index].boost_card, "has no more charges, removing it")
            else:
                new_card_count = int(active_player.active_boosts[index].boost_card["card_count"]) - 1
                active_player.active_boosts[index].boost_card["card_count"] = new_card_count
                data = create_boost_card_count_reduced_keyword("Boost has more than 1 charge, removing 1 from it, new charges left:", new_card_count)
                add_to_report(data)
                print("boost card with id:", active_player.active_boosts[index].boost_card, "has more than 1 charge, removing 1 from it, new charges left:",
                      new_card_count)

    if boost_to_remove_index is not None:
        active_player.active_boosts.pop(boost_to_remove_index)

    return active_player


def evaluate_combo_phase(battling_player1, battling_player2):
    print("----------------------------")
    print("Combo phase started")
    print("----------------------------")
    data = create_phase_started_report("combo_phase")
    add_to_report(data)

    # check for level1 combo
    battling_player1 = evaluate_combo_level_n(5, battling_player1)
    battling_player2 = evaluate_combo_level_n(5, battling_player2)
    # -----------------------------------------------------------------------------------------
    # check for level2 combo
    battling_player1 = evaluate_combo_level_n(4, battling_player1)
    battling_player2 = evaluate_combo_level_n(4, battling_player2)
    # -----------------------------------------------------------------------------------------
    # check for level3 combo
    battling_player1 = evaluate_combo_level_n(3, battling_player1)
    battling_player2 = evaluate_combo_level_n(3, battling_player2)
    # -----------------------------------------------------------------------------------------
    # check for level4 combo
    battling_player1 = evaluate_combo_level_n(2, battling_player1)
    battling_player2 = evaluate_combo_level_n(2, battling_player2)
    # -----------------------------------------------------------------------------------------
    # check for level5 combo
    battling_player1 = evaluate_combo_level_n(1, battling_player1)
    battling_player2 = evaluate_combo_level_n(1, battling_player2)

    data = create_phase_ended_report("combo_phase")
    add_to_report(data)

    print("----------------------------")
    print("Combo phase ended")
    print("----------------------------")

    return battling_player1, battling_player2


def check_for_combo_boosts(battling_player):
    combo_boosts = []
    if len(battling_player.active_boosts) > 0:
        for boost in battling_player.active_boosts:
            if boost.combo_boost is not None:
                combo_boosts.append(boost)
    return combo_boosts


def evaluate_combo_level_n(n, battling_player):

    if len(battling_player.player_combo_string) > (n + 1):
        combo_string = battling_player.player_combo_string[len(battling_player.player_combo_string) - (n + 2):]
        for combo in battling_player.player_combos:
            if combo["combo_code"] == combo_string:
                battling_player.player_combo_string = ""
                data= create_combo_found_report(battling_player.id,combo_string,n,combo)
                add_to_report(data)
                print(battling_player.player_dna,"found a level",n, "combo: ", combo_string, ". adding boost to player")
                if combo["target_type"] == "combo":
                    boost = Boost(combo)
                    data = create_boost_applied_report(battling_player.id,combo)
                    add_to_report(data)
                    battling_player.active_boosts.append(boost)
                else:
                    if combo["target_type"] == "player":
                        combo["card_timing"] = 0
                        combo["card_count"] = 1
                        if len(check_for_combo_boosts(battling_player))>0:
                            combo,battling_player = boost_combo(combo,battling_player)
                        combo_effect = ComboEffect(combo)
                        battling_player.combo_effects.append(combo_effect)
                        data=create_effect_applied_report(battling_player.id,combo)
                        add_to_report(data)
                    if combo["target_type"] == "card":
                        if len(check_for_combo_boosts(battling_player))>0:
                            combo,battling_player = boost_combo(combo,battling_player)
                        # TODO the reason why we are not boosting these things, because they would be way to op, I have to talk about this with someone
                        data = create_boost_applied_report(battling_player.id, combo)
                        add_to_report(data)
                        card_boost= Boost(combo)
                        battling_player.active_boosts.append(card_boost)
    return battling_player


def boost_combo(combo,battling_player):
    combo_boosts=check_for_combo_boosts(battling_player)
    first_boost=combo_boosts[0]

    # TODO someone told me each combo can only be boosted once, i dont remember who and why, so I might have to consult about this part and chane it

    if "shield" in combo:
        combo=boost_combo_with_parameter(combo,first_boost,battling_player,"shield")
    if "attack" in combo:
        combo = boost_combo_with_parameter(combo,first_boost, battling_player, "attack")
    if "life" in combo:
        combo = boost_combo_with_parameter(combo,first_boost, battling_player, "life")
    if "crit" in combo:
       combo=boost_combo_with_parameter(combo,first_boost, battling_player, "crit")

    battling_player = remove_boost_counter(battling_player, first_boost.unique_id)
    return combo, battling_player


def boost_combo_with_parameter(combo,first_boost,battling_player,parameter):
    # todo + type boost not supported yet, but there was no need for it with the current cards
    min_amount = combo[parameter]["amount"]
    if "extra" in combo[parameter]:
        max_amount = combo[parameter]["extra"]
        combo[parameter]["extra"] *= first_boost.combo_boost.amount
    combo[parameter]["amount"] *= first_boost.combo_boost.amount
    if "extra" in combo[parameter]:
        old_values= min_amount,"-",max_amount
        new_values=combo[parameter]["amount"], "-", combo[parameter]["extra"]
        data= create_combo_boosted_report(battling_player.id,parameter,"x",old_values,new_values,first_boost.combo_boost.amount)
        add_to_report(data)
        print(battling_player.player_dna, "original",parameter,"effect of", min_amount, "-", max_amount,
              "boosted to:", combo[parameter]["amount"], "-", combo[parameter]["extra"])
    else:
        old_values = min_amount
        new_values = combo[parameter]["amount"]
        data = create_combo_boosted_report(battling_player.id, parameter, "x", old_values, new_values,
                                           first_boost.combo_boost.amount)
        add_to_report(data)
        print(battling_player.player_dna, "original", parameter, "effect of", min_amount,"boosted to:", combo[parameter]["amount"])
    return combo


def evaluate_defense_phase(battling_player1, battling_player1_card_to_play, battling_player2,
                           battling_player2_card_to_play):
    print("----------------------------")
    print("Defense phase started")
    print("----------------------------")

    data = create_phase_started_report("defense_phase")
    add_to_report(data)
    # evaluate defense for player1
    evaluate_defense_phase_for_player(battling_player1, battling_player1_card_to_play)

    # evaluate defense for player2
    evaluate_defense_phase_for_player(battling_player2, battling_player2_card_to_play)

    data = create_phase_ended_report("defense_phase")
    add_to_report(data)
    print("----------------------------")
    print("Defense phase ended")
    print("----------------------------")

    return battling_player1, battling_player2


def evaluate_defense_phase_for_player(active_player, card_to_play):
    active_player = apply_defensive_combo_effects(active_player)

    if card_to_play is not None:
        if card_to_play["combo_sign"] == "D":
            life_boosts = []
            shield_boosts = []
            if len(active_player.active_boosts) > 0:
                for boost in active_player.active_boosts:
                    if boost.shield_boost is not None:
                        shield_boosts.append(boost.shield_boost)
                    if boost.life_boost is not None:
                        life_boosts.append(boost.life_boost)

            if len(shield_boosts) > 0 or len(life_boosts)>0:
                evaluate_defense_card(active_player, card_to_play, life_boosts, shield_boosts)
            else:
                evaluate_defense_card(active_player, card_to_play)


def apply_defensive_combo_effects(active_player):
    if len(active_player.combo_effects) > 0:
        for combo_effect in active_player.combo_effects:
            if combo_effect.shield_combo_effect is not None:
                active_player = apply_shield_combo_effect(active_player, combo_effect.shield_combo_effect)
            if combo_effect.life_combo_effect is not None:
                active_player = apply_life_combo_effect(active_player, combo_effect.life_combo_effect)
    return active_player


def apply_shield_combo_effect(active_player, shield_combo_effect):
    # TODO atm it does not support multiplication, ez to add but it was not needed cause there is no such card atm

    if shield_combo_effect.action_type == "+":
        shield_possible_values = []
        if shield_combo_effect.min_amount != shield_combo_effect.max_amount:
            for value in range(shield_combo_effect.min_amount, shield_combo_effect.max_amount+1):
                shield_possible_values.append(value)
            final_shield_value = random.choice(shield_possible_values)
        else:
            final_shield_value = shield_combo_effect.min_amount
        active_player.player_shield += final_shield_value
        data = create_defense_effect_activated_report(active_player.id,"shield",final_shield_value,active_player.player_shield)
        add_to_report(data)
        print(active_player.player_dna, "gained", final_shield_value, "shield from a combo effect, his shield now is:",
              active_player.player_shield)

    return active_player


def apply_life_combo_effect(active_player, life_combo_effect):
    # TODO atm it does not support multiplication, ez to add but it was not need so i did not do that

    if life_combo_effect.action_type == "+":
        life_possible_values = []
        if life_combo_effect.min_amount != life_combo_effect.max_amount:
            for value in range(life_combo_effect.min_amount, life_combo_effect.max_amount+1):
                life_possible_values.append(value)
            final_life_value = random.choice(life_possible_values)
        else:
            final_life_value = life_combo_effect.min_amount

        active_player.player_health += final_life_value
        if active_player.player_health > active_player.player_max_health:
            active_player.player_health = active_player.player_max_health

        data = create_defense_effect_activated_report(active_player.id, "life", final_life_value,active_player.player_health)
        add_to_report(data)
        print(active_player.player_dna, "gained", final_life_value, "life from a combo effect, his HP now is:",
              active_player.player_health)

    return active_player


def evaluate_defense_card(active_player, card_to_play, life_boosts=[], shield_boosts=[]):
    # check if the card is of type shield
    if "shield" in card_to_play:
        # calculating final shield value
        evaluate_shield_card(active_player,card_to_play,shield_boosts)

    else:
        if "life" in card_to_play:
            # calculating final life value
            evaluate_life_card(active_player,card_to_play,life_boosts)


def evaluate_life_card(active_player,card_to_play,life_boosts):
    min_life_value = card_to_play["life"]["amount"]
    max_life_value = card_to_play["life"]["extra"]
    random_life_array = []
    if min_life_value == max_life_value:
        final_life_value = min_life_value
    else:
        for life_value in range(min_life_value, max_life_value+1):
            random_life_array.append(life_value)
        final_life_value = random.choice(random_life_array)

    data = create_defense_played_report(active_player.id, "life", min_life_value, max_life_value,
                                        final_life_value)
    add_to_report(data)
    # evaluating life boost
    if len(life_boosts) > 0:
        final_life_value, active_player = evaluate_life_boosts(final_life_value, active_player, life_boosts)

    # apply life to player
    active_player.player_health += final_life_value

    # capping the hp
    if active_player.player_health > active_player.player_max_health:
        health_surplus = active_player.player_health - active_player.player_max_health
        active_player.player_health = active_player.player_max_health
        final_life_value -= health_surplus
    data = create_defense_applied_report(active_player.id, "life", final_life_value,active_player.player_health,active_player.player_shield)
    add_to_report(data)
    print(active_player.player_dna, "gained ", final_life_value, " life, his total is now: ",
          active_player.player_health)
    return active_player


def evaluate_shield_card(active_player,card_to_play,shield_boosts):
    min_shield_value = card_to_play["shield"]["amount"]
    max_shield_value = card_to_play["shield"]["extra"]
    random_shield_array = []
    if min_shield_value == max_shield_value:
        final_shield_value = min_shield_value
    else:
        for shield_value in range(min_shield_value, max_shield_value+1):
            random_shield_array.append(shield_value)
        final_shield_value = random.choice(random_shield_array)

    data = create_defense_played_report(active_player.id, "shield", min_shield_value, max_shield_value,final_shield_value)
    add_to_report(data)
    # evaluating shield boost
    if len(shield_boosts) > 0:
        final_shield_value, active_player = evaluate_shield_boosts(final_shield_value, active_player, shield_boosts)
    # apply shield to player
    active_player.player_shield += final_shield_value
    data = create_defense_applied_report(active_player.id, "shield", final_shield_value, active_player.player_health,
                                         active_player.player_shield)
    add_to_report(data)
    print(active_player.player_dna, "gained ", final_shield_value, "shield, his current total is: ",
          active_player.player_shield)
    return  active_player


def evaluate_life_boosts(final_value, active_player, active_life_boosts):
    # evaluate the life boosts that add flat value
    starting_value=copy.deepcopy(final_value)
    for life_boost in active_life_boosts:
        if life_boost.action_type == "+":
            final_value += life_boost.amount
            data = create_defense_boosted_report(active_player.id, "life", life_boost.amount, "+", starting_value, final_value)
            add_to_report(data)
            active_player = remove_boost_counter(active_player, life_boost.unique_id)
            print("Life boosted by +", life_boost.amount, " new final_life_value is: ", final_value)
    # evaluate the life boosts that multiply the value
    for life_boost in active_life_boosts:
        if life_boost.action_type == "x":
            final_value = final_value * life_boost.amount
            data = create_defense_boosted_report(active_player.id, "life", life_boost.amount, "x", starting_value,
                                                 final_value)
            add_to_report(data)
            active_player = remove_boost_counter(active_player, life_boost.unique_id)
            print("Life boosted by x", life_boost.amount, " new final_life_value is: ", final_value)
            return final_value, active_player

    return final_value, active_player


def evaluate_shield_boosts(final_value, active_player, active_shield_boosts):
    # evaluate the shield boosts that add flat value
    # The reason why this is a little different than the life boosts,
    # is because there are combos that create boosts with non- flat value
    starting_value= copy.deepcopy(final_value)
    for shield_boost in active_shield_boosts:
        if shield_boost.action_type == "+":
            if shield_boost.extra is not None and shield_boost.amount == shield_boost.extra:
                boost_values=[]
                for number in range(shield_boost.amount, shield_boost.extra+1):
                    boost_values.append(number)
                    final_boost_amount=random.choice(boost_values)
                final_value += final_boost_amount
            else:
                final_boost_amount=shield_boost.amount
                final_value += final_boost_amount

            data = create_defense_boosted_report(active_player.id, "shield", final_boost_amount, "+", starting_value,
                                                 final_value)
            add_to_report(data)
            print(active_player.player_dna, "shield boosted by +", shield_boost.amount, " new final_shield_value is: ",
                  final_value)
            active_player = remove_boost_counter(active_player, shield_boost.unique_id)
    # evaluate the shield boosts that multiply the value
    for shield_boost in active_shield_boosts:
        if shield_boost.action_type == "x":
            final_value = final_value * shield_boost.amount
            data = create_defense_boosted_report(active_player.id, "shield", shield_boost.amount, "x", starting_value,
                                                 final_value)
            add_to_report(data)
            print("Shield boosted by x", shield_boost.amount, " new final_shield_value is: ", final_value)
            active_player = remove_boost_counter(active_player, shield_boost.unique_id)
            return final_value, active_player

    return final_value, active_player


def evaluate_attack_phase(battling_player1, battling_player1_card_to_play, battling_player2,
                          battling_player2_card_to_play):
    print("----------------------------")
    print("Attack phase started")
    print("----------------------------")

    data = create_phase_started_report("attack_phase")
    add_to_report(data)

    # evaluate attack for player1
    evaluate_attack_phase_for_player(battling_player1, battling_player2, battling_player1_card_to_play)

    # evaluate attack for player2
    evaluate_attack_phase_for_player(battling_player2, battling_player1, battling_player2_card_to_play)

    data = create_phase_ended_report("attack_phase")
    add_to_report(data)

    print("----------------------------")
    print("Attack phase ended")
    print("----------------------------")

    return battling_player1, battling_player2


def apply_offensive_combo_effects(active_player, target_player):
    if len(active_player.combo_effects) > 0:
        for combo_effect in active_player.combo_effects:
            if combo_effect.attack_combo_effect is not None:
                target_player = apply_attack_combo_effect(active_player,target_player, combo_effect.attack_combo_effect)
    return target_player


def apply_attack_combo_effect(attacking_player,defending_player, attack_combo_effect):
    if attack_combo_effect.action_type == "+":
        attack_possible_values = []
        if attack_combo_effect.min_amount != attack_combo_effect.max_amount:
            for value in range(attack_combo_effect.min_amount, attack_combo_effect.max_amount+1):
                attack_possible_values.append(value)
            final_attack_value = random.choice(attack_possible_values)
        else:
            final_attack_value = attack_combo_effect.min_amount
        data = create_attack_effect_activated_report(attacking_player.id,attack_combo_effect.min_amount,attack_combo_effect.max_amount,final_attack_value)
        add_to_report(data)
        print("Dealing damage from combo source")
        defending_player = deal_damage(attacking_player,final_attack_value, defending_player, False)
    return defending_player


def evaluate_attack_phase_for_player(attacking_player, defending_player, card_to_play):
    defending_player = apply_offensive_combo_effects(attacking_player, defending_player)
    attacking_player = check_for_crit_boost(attacking_player)

    if card_to_play is not None:
        if card_to_play["combo_sign"] == "A":
            attacking_player, is_piercing = check_for_pierce(attacking_player)
            attack_boosts = []
            if len(attacking_player.active_boosts) > 0:
                for boost in attacking_player.active_boosts:
                    if boost.attack_boost is not None:
                        attack_boosts.append(boost.attack_boost)
            if len(attack_boosts) > 0:
                evaluate_attack_card(attacking_player, defending_player, card_to_play,
                                     attacking_player.player_crit_chance, is_piercing, attack_boosts)
            else:
                evaluate_attack_card(attacking_player, defending_player, card_to_play,
                                     attacking_player.player_crit_chance, is_piercing)


def check_for_pierce(attacking_player):
    if len(attacking_player.active_boosts) > 0:
        for boost in attacking_player.active_boosts:
            if boost.special_boost is not None:
                if boost.special_boost.special == "add pierce":
                    data = create_pierce_applied_report(attacking_player.id,True)
                    add_to_report(data)
                    print(attacking_player.player_dna, " has pierce")
                    remove_boost_counter(attacking_player, boost.unique_id)
                    return attacking_player, True
    return attacking_player, False


def check_for_crit_boost(attacking_player):
    if len(attacking_player.active_boosts) > 0:
        for boost in attacking_player.active_boosts:
            if boost.crit_boost is not None:
                if boost.crit_boost.action_type == "+":
                    base_crit_chance = attacking_player.player_crit_chance
                    attacking_player.player_crit_chance += boost.crit_boost.amount
                    if attacking_player.player_crit_chance > 100:
                        attacking_player.player_crit_chance = 100
                    data = create_crit_boosted(attacking_player.id,boost.crit_boost.action_type,boost.crit_boost.amount,base_crit_chance,attacking_player.player_crit_chance)
                    add_to_report(data)
                    print(attacking_player.player_dna, "boosted crit chance from", base_crit_chance, "to",
                          attacking_player.player_crit_chance)
                    attacking_player = remove_boost_counter(attacking_player, boost.crit_boost.unique_id)


    return attacking_player


def evaluate_attack_card(attacking_player, defending_player, card_to_play, attacking_player_crit_ratio, is_piercing,
                         active_boosts=[]):
    # calculating the final damage based on crit ration and attack range
    min_attack_damage = card_to_play["attack"]["amount"]
    max_attack_damage = card_to_play["attack"]["extra"]
    print(attacking_player.player_dna, "Attack card played with:", min_attack_damage, "-", max_attack_damage,
          "and crit chance:", attacking_player_crit_ratio)
    if attacking_player_crit_ratio == 100:
        final_damage = max_attack_damage
        print(attacking_player.player_dna, "Critical hit")
        print(attacking_player.player_dna, "The final damage value is:", final_damage)
        is_crit=1
    else:
        random_array = [0] * 100
        for i in range(attacking_player_crit_ratio+1):
            random_array[i] = 1
        is_crit = random.choice(random_array)
        if is_crit:
            final_damage = max_attack_damage
            print(attacking_player.player_dna, "Critical hit")
            print(attacking_player.player_dna, "The final damage value is:", final_damage)
        else:
            damage_array = []
            if min_attack_damage == max_attack_damage:
                final_damage=min_attack_damage
            else:
                for d in range(min_attack_damage, max_attack_damage+1):
                    damage_array.append(d)
                final_damage = random.choice(damage_array)
            print(attacking_player.player_dna, "The final damage value is:", final_damage)

    data = create_attack_played_report(attacking_player.id, min_attack_damage, max_attack_damage,
                                       attacking_player_crit_ratio, is_crit, final_damage)
    add_to_report(data)
    # evaluating attack boost
    if len(active_boosts) > 0:
        final_damage, attacking_player = evaluate_attack_boosts(final_damage, attacking_player, active_boosts)

    defending_player = deal_damage(attacking_player,final_damage, defending_player, is_piercing)


def evaluate_attack_boosts(final_damage, attacking_player, active_boosts):
    # evaluate flat bonus damage
    starting_value=copy.deepcopy(final_damage)
    for attack_boost in active_boosts:
        if attack_boost.action_type == "+":
            if attack_boost.extra is not None and attack_boost.amount == attack_boost.extra:
                boost_values = []
                for number in range(attack_boost.amount, attack_boost.extra+1):
                    boost_values.append(number)
                    final_boost_amount=random.choice(boost_values)
                final_damage += final_boost_amount
            else:
                final_boost_amount = attack_boost.amount
                final_damage += final_boost_amount

            data = create_attack_boosted_report(attacking_player.id,"+",final_boost_amount,starting_value,final_damage)
            add_to_report(data)

            print(attacking_player.player_dna, "attack boosted by +", attack_boost.amount, " new attack value is: ",
                  final_damage)
            attacking_player = remove_boost_counter(attacking_player, attack_boost.unique_id)

    # evaluate multiplayer bonus damage
    for attack_boost in active_boosts:
        if attack_boost.action_type == "x":
            final_damage *= attack_boost.amount

            data = create_attack_boosted_report(attacking_player.id, "x", attack_boost.amount, starting_value,
                                                final_damage)
            add_to_report(data)

            print(attacking_player.player_dna, "boosted base attack of", starting_value, "by x", attack_boost.amount,
                  " new attack_final_value is: ", final_damage)
            attacking_player = remove_boost_counter(attacking_player, attack_boost.unique_id)
            return final_damage, attacking_player

    return final_damage, attacking_player


def deal_damage(attacking_player,final_damage, defending_player, is_piercing):
    # deal pierce damage (ignore armor)
    if is_piercing:
        defending_player.player_health -= final_damage
        print(defending_player.player_dna, "took", final_damage,
              "pierce damage, his HP now is:", defending_player.player_health)
    else:
        # taking player shield into consideration
        if defending_player.player_shield > 0:
            if defending_player.player_shield >= final_damage:
                damage_blocked = final_damage
                defending_player.player_shield -= final_damage
                data = create_damage_dealt_report(attacking_player.id, defending_player.id, damage_blocked, 0,
                                                  defending_player.player_health, defending_player.player_shield)
                add_to_report(data)
                damages_dealt[attacking_player.id] += final_damage
                print(defending_player.player_dna, "blocked ", damage_blocked, "and  took ", 0,
                      "damage, his HP now is:",
                      defending_player.player_health, "and shield: ", defending_player.player_shield)
            else:
                damage_blocked = defending_player.player_shield
                final_damage -= defending_player.player_shield
                defending_player.player_shield = 0
                defending_player.player_health -= final_damage
                data = create_damage_dealt_report(attacking_player.id, defending_player.id, damage_blocked, final_damage,
                                                  defending_player.player_health, defending_player.player_shield)
                add_to_report(data)
                damages_dealt[attacking_player.id] +=final_damage
                print(defending_player.player_dna, "blocked ", damage_blocked, "and  took ", final_damage,
                      " damage, his HP now is: ", defending_player.player_health, "and shield: ",
                      defending_player.player_shield)
        else:
            defending_player.player_health -= final_damage
            data = create_damage_dealt_report(attacking_player.id,defending_player.id,0,final_damage,defending_player.player_health, defending_player.player_shield)
            add_to_report(data)
            damages_dealt[attacking_player.id] += final_damage
            print(defending_player.player_dna, "took ", final_damage, "damage, his HP now is: ",
                  defending_player.player_health, "and shield: ", defending_player.player_shield)

    return defending_player


def determine_winner(battling_player1, battling_player2):
    if battling_player1.player_health > battling_player2.player_health:
        return battling_player1,battling_player2
    elif battling_player1.player_health < battling_player2.player_health:

        return battling_player2,battling_player1
    elif battling_player1.player_shield > battling_player2.player_shield:

        return battling_player1,battling_player2
    elif battling_player1.player_shield < battling_player2.player_shield:

        return battling_player2,battling_player1
    elif battling_player1.id>battling_player2.id:

        return battling_player1,battling_player2
    else:

        return battling_player2,battling_player1


# this is what you run
workbook = load_workbook(filename=rules_filename)
layers = load_layers(workbook["layers"])

playing_cards = load_cards(workbook["cards"])
combos = load_combos(workbook["combos"])
player_decks=[]
# load cards
with open(cards_filename, encoding='utf-8') as json_file:
    ether_cards = json.load(json_file)
    json_file.close()


def display_player_info(player):
        print("-------------------------------------------------")
        print("Player id:",player.params["card_id"])
        print("Deck limit:",player.params["deck_limit"])
        print("Deck max cost: 200")
        print("Character type:", player.params["character_type"])
        print("Combo group:",player.params["combo_group"])
        print("Player max health:", player.params["health"])
        print("Player base crit chance:",player.params["crit"])
        print("Player dna:",player.params["dna"])

        print("-------------------------------------------------")


def find_ether_card(id):
    for ec in ether_cards:
        if ec["id"] == id:
            return ec
    return "Not found"


# assign the stats from cards.json to the players
def get_ether_card(id):
    ether_card=find_ether_card(id)
    if ether_card != "Not fund":
        return ether_card


def transform_deck_code(deck_code):
    deck=[]
    for card_code in deck_code:
        card=playing_cards[card_code["id"]]
        deck.append(card)
    return deck


def get_deck(id):

    filename = str(id)+".json"
    full_path= os.path.join(players_folder_path,filename)
    with open(full_path, encoding='utf-8') as json_file:
        player_data = json.load(json_file)
        json_file.close()
    deck_code= player_data["deck"]
    return transform_deck_code(deck_code)


def simulate_battle(match_unique_id,player1_id,player2_id):
    #old_stdout = sys.stdout
    #log_report_identifier=match_unique_id+"_"+str(player1_id)+"_"+str(player2_id)+".log"
    #full_path = os.path.join(log_reports_folder_path, log_report_identifier)
    #log_file = open(full_path, "w")
    #sys.stdout = log_file

    # TODO JUST FOR TESTING
    ether_card1=get_ether_card(player1_id)
    ether_card2=get_ether_card(player2_id)

    player1 = Player(ether_card1, layers, combos, playing_cards)
    player2 = Player(ether_card2, layers, combos, playing_cards)

    display_player_info(player1)
    display_player_info(player2)

    player1_deck = get_deck(player1_id)
    player2_deck = get_deck(player2_id)

    player1.validate_and_assign_deck(player1_deck)
    #player1.generate_random_deck()
    player2.validate_and_assign_deck(player2_deck)
    #player2.generate_random_deck()
    player1_score = 0
    player2_score = 0
    player1_id = player1.params["card_id"]
    player2_id = player2.params["card_id"]
    damages_dealt[player1.params["card_id"]]=0
    damages_dealt[player2.params["card_id"]]=0
    total_healths[player1.params["card_id"]]=0
    total_healths[player2.params["card_id"]]=0

    for i in range(9):
        (round_winner,round_loser) = battle(player1, player2)
        print("The winner is: ", round_winner.player_dna)
        if round_winner.id == player1_id:
            player1_score += 1
        elif round_winner.id == player2_id:
            player2_score += 1
        print("The score is:", player1_score, "-", player2_score)
        total_healths[round_winner.id] += round_winner.player_health
        total_healths[round_loser.id] += round_loser.player_health
        data = create_end_of_game_report(round_winner.id, round_loser.id)
        add_to_report(data)
    damages_dealt_json={player1_id:damages_dealt[player1_id],player2_id:damages_dealt[player2_id]}
    total_healths_json = {player1_id:total_healths[player1_id], player2_id:total_healths[player2_id]}
    if player1_score> player2_score:
        data= create_end_of_series_report(player1.params["card_id"],player2.params["card_id"],damages_dealt_json,total_healths_json,
                                       {player1.params["card_id"]:player1_score,player2.params["card_id"]: player2_score})
    else:
        data = create_end_of_series_report(player2.params["card_id"], player1.params["card_id"], damages_dealt_json,
                                           total_healths_json,
                                           {player1.params["card_id"]: player1_score,
                                            player2.params["card_id"]: player2_score})
    add_to_report(data)
    damages_dealt.clear()
    total_healths.clear()

    report_identifier=match_unique_id+"_"+str(player1_id)+"_"+str(player2_id)+".json"
    full_path = os.path.join(report_folder_path,report_identifier)
    last_report = copy.deepcopy(series_report[-1])
    with open (full_path, 'a') as outfile :
        outfile.truncate(0)
        outfile.write(json.dumps(series_report))
        series_report.clear()
        outfile.close()

    #sys.stdout = old_stdout
    #log_file.close()
    return last_report


for i in range (1):
    match_unique_id = str(uuid.uuid4())
    simulate_battle(match_unique_id,311 ,4660)
